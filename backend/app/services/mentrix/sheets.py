"""Mentrix Sheets — chat-driven workbook JSON + XLSX. Formulas stored, never executed."""

from __future__ import annotations

import io
import json
import re
import zipfile
from pathlib import Path
from typing import Any

from app.infrastructure.allowed_paths import path_under_allowed_roots

MAX_UPLOAD = 8 * 1024 * 1024
MAX_MEMBERS = 200
MAX_UNCOMPRESSED = 40 * 1024 * 1024
_CELL = re.compile(r"^[A-Z]{1,3}[1-9][0-9]{0,4}$")


def empty_workbook() -> dict[str, Any]:
    return {"sheets": [{"name": "Sheet1", "cells": {}}]}


def normalize_workbook(raw: Any) -> dict[str, Any]:
    sheets_in = []
    if isinstance(raw, dict):
        sheets_in = list(raw.get("sheets") or [])
    out = []
    for sheet in sheets_in[:8]:
        if not isinstance(sheet, dict):
            continue
        name = str(sheet.get("name") or "Sheet")[:40]
        cells_in = sheet.get("cells") if isinstance(sheet.get("cells"), dict) else {}
        cells: dict[str, dict[str, str]] = {}
        for key, val in list(cells_in.items())[:400]:
            addr = str(key).upper().replace("$", "")
            if not _CELL.match(addr):
                continue
            if isinstance(val, dict):
                cells[addr] = {
                    "v": str(val.get("v") if val.get("v") is not None else "")[:2000],
                    "f": str(val.get("f") or "")[:500],
                }
            else:
                cells[addr] = {"v": str(val)[:2000], "f": ""}
        out.append({"name": name, "cells": cells})
    if not out:
        return empty_workbook()
    return {"sheets": out}


def generate_workbook(prompt: str, *, project_id: int | None = None) -> dict[str, Any]:
    from app.adapters.llm.openai_compat import get_openai_compat_client, openai_compat_available, mentrix_llm_chat_model
    from app.services.coding_engine.agent_context import compose_coding_agent_context

    if not openai_compat_available():
        raise ValueError("llm_offline")
    pack = ""
    if project_id:
        pack = compose_coding_agent_context(goal=prompt, project_id=project_id, max_chars=800)
    client = get_openai_compat_client()
    resp = client.chat.completions.create(
        model=mentrix_llm_chat_model(),
        messages=[
            {
                "role": "system",
                "content": (
                    "Return JSON only: {\"sheets\":[{\"name\":\"Sheet1\",\"cells\":{\"A1\":{\"v\":\"text\",\"f\":\"\"}}}]}. "
                    "Small grids only. Never invent finance figures. Store formulas as text in f; do not evaluate."
                ),
            },
            {"role": "user", "content": f"{prompt[:2000]}\n\n{pack}"},
        ],
        max_tokens=1200,
        temperature=0.2,
    )
    raw = resp.choices[0].message.content or "{}"
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        start, end = raw.find("{"), raw.rfind("}")
        data = json.loads(raw[start : end + 1]) if start >= 0 and end > start else {}
    return normalize_workbook(data)


def workbook_to_xlsx(workbook: dict[str, Any]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for sheet in normalize_workbook(workbook)["sheets"]:
        ws = wb.active if first else wb.create_sheet(title=sheet["name"][:31])
        if first:
            ws.title = sheet["name"][:31]
            first = False
        for addr, cell in sheet["cells"].items():
            formula = (cell.get("f") or "").strip()
            ws[addr] = formula if formula.startswith("=") else cell.get("v") or ""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _assert_safe_xlsx(data: bytes) -> None:
    if len(data) > MAX_UPLOAD:
        raise ValueError("file_too_large")
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as exc:
        raise ValueError("not_xlsx") from exc
    if len(zf.infolist()) > MAX_MEMBERS:
        raise ValueError("zip_bomb")
    total = 0
    for info in zf.infolist():
        if info.file_size > MAX_UNCOMPRESSED or info.file_size < 0:
            raise ValueError("zip_bomb")
        total += info.file_size
        if total > MAX_UNCOMPRESSED:
            raise ValueError("zip_bomb")


def workbook_from_xlsx(data: bytes) -> dict[str, Any]:
    _assert_safe_xlsx(data)
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=False, read_only=True)
    sheets = []
    for ws in wb.worksheets[:8]:
        cells: dict[str, dict[str, str]] = {}
        for row in ws.iter_rows(max_row=40, max_col=20):
            for cell in row:
                if cell.value is None:
                    continue
                addr = cell.coordinate
                val = str(cell.value)
                formula = val if val.startswith("=") else ""
                shown = "" if formula else val
                cells[addr] = {"v": shown[:2000], "f": formula[:500]}
        sheets.append({"name": ws.title[:40], "cells": cells})
    wb.close()
    return normalize_workbook({"sheets": sheets})


def resolve_workbook_path(raw: str) -> Path:
    text = (raw or "").strip()
    if not text:
        raise ValueError("path_required")
    if ".." in text.replace("\\", "/").split("/"):
        raise ValueError("path_escape")
    path = Path(text).expanduser()
    try:
        resolved = path_under_allowed_roots(str(path))
    except ValueError as exc:
        raise ValueError("path_escape") from exc
    parts = {p.lower() for p in resolved.parts}
    if ".zect" not in parts or "workbooks" not in parts:
        raise ValueError("must_be_under_zect_workbooks")
    if resolved.suffix.lower() not in {".json", ".xlsx"}:
        raise ValueError("must_be_json_or_xlsx")
    return resolved


def save_workbook(path: Path, workbook: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(normalize_workbook(workbook), indent=2), encoding="utf-8")
